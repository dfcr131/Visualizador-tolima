# app.py
import streamlit as st
import pandas as pd
from pathlib import Path
import plotly.express as px
import folium
from streamlit_folium import st_folium
import requests
import os
from PIL import Image, ImageOps
# ================== CONFIG BÁSICA ==================
st.set_page_config(
    page_title="Información cualitativa departamental",
    page_icon="data/ubicacion.png",
    layout="wide"
)
# =========================================================
# 🌿 ESTILOS ELEGANTES UNIFICADOS + KPI CARDS
# =========================================================
st.markdown("""
<style>

/* =========================================================
   🎨 ESTILO GLOBAL ELEGANTE - VERDE MENTA
   ========================================================= */

/* Fuente general */
html, body, [class*="css"] {
    font-family: 'Poppins', sans-serif !important;
    background-color: #f9fdfb !important;
}

/* =========================================================
   🧭 SIDEBAR
   ========================================================= */
div[data-testid="stSidebar"] {
    background-color: #e9fdf9 !important;
    border-right: 3px solid #3fb4a1 !important;
    box-shadow: 4px 0 8px rgba(63,180,161,0.1);
    padding: 25px 20px !important;
    border-top-right-radius: 18px !important;
}

/* Títulos de los filtros */
div[data-testid="stSidebar"] label {
    color: #106c5d !important;
    font-weight: 600 !important;
    font-size: 15px !important;
}

/* Selects y entradas */
div[data-baseweb="select"], input, textarea {
    background-color: #ffffff !important;
    color: #154734 !important;
    border-radius: 10px !important;
    border: 1.8px solid #3fb4a1 !important;
    transition: all 0.3s ease;
    box-shadow: 0px 2px 5px rgba(63,180,161,0.1);
}

div[data-baseweb="select"]:hover, input:hover, textarea:hover {
    border-color: #2d9b8d !important;
    box-shadow: 0 0 8px rgba(63,180,161,0.25);
}

/* Botón de limpiar filtros */
button[kind="secondary"] {
    background-color: #3fb4a1 !important;
    color: white !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    border: none !important;
    transition: all 0.3s ease-in-out;
}
button[kind="secondary"]:hover {
    background-color: #2d9b8d !important;
    box-shadow: 0 4px 10px rgba(63,180,161,0.3);
}

/* Texto general del sidebar */
div[data-testid="stSidebar"] p, div[data-testid="stSidebar"] span {
    color: #106c5d !important;
}

/* =========================================================
   🌿 CUERPO PRINCIPAL
   ========================================================= */
.main {
    background-color: #f9fdfb !important;
}

/* Títulos */
h1, h2, h3 {
    color: #106c5d !important;
    font-weight: 700 !important;
    text-shadow: 0px 1px 2px rgba(63,180,161,0.15);
}

h2::after {
    content: "";
    display: block;
    width: 80px;
    height: 4px;
    background-color: #3fb4a1;
    margin-top: 6px;
    border-radius: 4px;
}

/* =========================================================
   📊 TARJETAS DE MÉTRICAS
   ========================================================= */
div[data-testid="stMetricValue"] {
    color: #106c5d !important;
    font-weight: 700 !important;
}

div[data-testid="stMetricLabel"] {
    color: #3fb4a1 !important;
    font-weight: 600 !important;
}

/* Contenedor de métricas */
div.css-1ht1j8u, div.css-12w0qpk, div.css-1r6slb0 {
    background-color: #ffffff !important;
    border: 2px solid #3fb4a1 !important;
    border-radius: 14px !important;
    box-shadow: 0px 4px 10px rgba(63,180,161,0.15) !important;
    padding: 18px !important;
    transition: transform 0.3s ease, box-shadow 0.3s ease;
}

div.css-1ht1j8u:hover, div.css-12w0qpk:hover, div.css-1r6slb0:hover {
    transform: translateY(-4px);
    box-shadow: 0px 6px 14px rgba(63,180,161,0.3) !important;
}

/* =========================================================
   💎 TARJETAS KPI PERSONALIZADAS
   ========================================================= */
.kpi-card {
    border-radius: 20px;
    padding: 20px;
    background-color: #3fb4a1;
    color: white;
    box-shadow: 0 4px 12px rgba(63,180,161,0.4);
    text-align: center;
    transition: all 0.3s ease;
    border: 3px solid #3fb4a1;
    position: relative;
    overflow: hidden;
}

/* Borde decorativo dinámico */
.kpi-card::before {
    content: "";
    position: absolute;
    inset: 0;
    border-radius: 18px;
    border: 2px dashed rgba(255,255,255,0.25);
    pointer-events: none;
}

/* Hover con brillo */
.kpi-card:hover {
    transform: translateY(-4px);
    box-shadow: 0 6px 18px rgba(63,180,161,0.5);
}

/* Título */
.kpi-title {
    font-size: 14px;
    opacity: 0.95;
    font-weight: 500;
    color: #e0f8f3;
    margin-bottom: 4px;
}

/* Valor */
.kpi-value {
    font-size: 32px;
    font-weight: 700;
    color: #ffffff;
    margin: 0;
}

/* Línea decorativa */
.kpi-value::after {
    content: "";
    display: block;
    width: 40%;
    height: 3px;
    margin: 6px auto 0 auto;
    background-color: white;
    border-radius: 2px;
}

/* =========================================================
   📱 Adaptativo
   ========================================================= */
@media (max-width: 768px) {
    .kpi-card {
        padding: 14px;
    }
    .kpi-value {
        font-size: 24px;
    }
}

</style>
""", unsafe_allow_html=True)

# ================== ENCABEZADO ==================
import streamlit as st

# ======= ENCABEZADO PRINCIPAL =======
col_title, col_extra = st.columns([5, 2], vertical_alignment="center")

with col_title:
    st.markdown("""
        <div style="
            text-align: left;
            padding: 20px 10px;
        ">
            <h1 style="
                color: #106c5d;
                font-weight: 700;
                font-size: 2.2rem;
                margin-bottom: 0.4rem;
                text-shadow: 0px 1px 2px rgba(63,180,161,0.15);
            ">
                Bienvenido a la Información Cualitativa Departamental
            </h1>
            <p style="
                color: #3fb4a1;
                font-size: 1.05rem;
                font-weight: 500;
                margin-top: 0.2rem;
            ">
                Levantamiento de información con instrumento de Web Scraping y Análisis de Redes Sociales
            </p>
        </div>
    """, unsafe_allow_html=True)

with col_extra:
    st.markdown(
        """
        <div style="text-align: right; padding-right: 15px;">
        """,
        unsafe_allow_html=True
    )
    try:
        st.image("data/fontur_logo.png", width=1400)
    except Exception:
        st.markdown("<p style='color:#3fb4a1;'>Logo no disponible</p>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

# ======= DIVISOR DECORATIVO =======
st.markdown("""
<hr style="
    border: none;
    height: 3px;
    background-color: #3fb4a1;
    border-radius: 3px;
    margin-top: 10px;
    margin-bottom: 25px;
">
""", unsafe_allow_html=True)


# ================== GALERÍA DE IMÁGENES ==================
# ================== GALERÍA DE IMÁGENES ==================
carpeta_imagenes = "./data/imagenes"

if os.path.exists(carpeta_imagenes):
    imagenes = [
        os.path.join(carpeta_imagenes, img)
        for img in os.listdir(carpeta_imagenes)
        if img.lower().endswith(('.png', '.jpg', '.jpeg', '.gif'))
    ]

    if imagenes:
        st.markdown("<h4 style='text-align:center; color:#1e5631;'>Galería de Imágenes</h4>", unsafe_allow_html=True)
        cols = st.columns(min(4, len(imagenes)))  # máximo 4 columnas

        for i, img_path in enumerate(imagenes):
            with cols[i % len(cols)]:
                try:
                    abs_path = os.path.abspath(img_path)
                    img = Image.open(abs_path)
                    img = ImageOps.fit(img, (300, 180), Image.LANCZOS, centering=(0.5, 0.5))
                    st.image(img, use_container_width=False)
                except Exception as e:
                    st.warning(f"No se pudo cargar: {img_path}\nError: {e}")

        # 🔽 Texto después de las imágenes
        st.markdown(
            """
            <div style='text-align:center; margin-top:20px; font-size:16px; color:#444;'>
                 Las imágenes presentadas en esta galería son de carácter ilustrativo y fueron obtenidas de fuentes públicas en Internet.
            </div>
            """,
            unsafe_allow_html=True
        )

    else:
        st.info("No se encontraron imágenes en la carpeta.")
else:
    st.warning(f"La carpeta '{carpeta_imagenes}' no existe.")

# ================== CARGA DE DATOS (con hipervínculos) ==================
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import streamlit as st

DEFAULT_FILE = Path("data") / "consolidado_turismo LOTE 1 final.xlsx"

# ✅ Lista de hojas válidas
VALID_SHEETS = ["Cod Tol", "Cod Putumayo", "Cod Huila", "Cod Caquetá"]

@st.cache_data(show_spinner=False)
def read_excel_all(file: Path, valid_sheets):
    """
    Lee todas las hojas de un Excel e incluye tanto el texto visible como los hipervínculos.
    Si una celda contiene un enlace, se crea una columna adicional con el sufijo '_URL'.
    """
    wb = load_workbook(file, data_only=True)
    df_list = []

    for sheet in valid_sheets:
        if sheet not in wb.sheetnames:
            st.warning(f"⚠️ La hoja '{sheet}' no existe en el archivo.")
            continue

        ws = wb[sheet]

        # Obtener encabezados
        headers = [cell.value for cell in ws[1]]
        if not headers:
            st.warning(f"⚠️ La hoja '{sheet}' no tiene encabezados válidos.")
            continue

        data = []
        for row in ws.iter_rows(min_row=2, max_col=len(headers)):
            values = []
            for cell in row:
                text = cell.value
                url = cell.hyperlink.target if cell.hyperlink else None
                values.append((text, url))
            data.append(values)

        # Crear diccionario: texto + posibles URLs
        df_dict = {}
        for j, header in enumerate(headers):
            # Texto visible
            df_dict[header] = [data[i][j][0] for i in range(len(data))]
            # Enlace si existe
            urls = [data[i][j][1] for i in range(len(data))]
            if any(urls):
                df_dict[f"{header}_URL"] = urls

        df_temp = pd.DataFrame(df_dict)
        df_list.append(df_temp)

    if not df_list:
        st.error("No se pudieron leer las hojas especificadas.")
        st.stop()

    df = pd.concat(df_list, ignore_index=True)
    return df


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.str.strip()
    aliases = {
        "DEPARTAMENTO": "Departamento",
        "Departamento ": "Departamento",
        "MUNICIPIO": "Municipio",
        "ENFOQUE TURÍSTICO": "Enfoque Turístico",
        "ENFOQUE TURISTICO": "Enfoque Turístico",
        "DESCRIPCIÓN": "Descripción",
        "Descripcion": "Descripción",
        "TITULO": "Título",
        "ACTOR": "Actor",
        "SECTOR": "Sector",
        "ASPECTO": "Aspecto",
        "NOMBRE": "Nombre",
    }
    df.rename(columns={k: v for k, v in aliases.items() if k in df.columns}, inplace=True)
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    return df


# ================== VALIDACIÓN Y EJECUCIÓN ==================
if not DEFAULT_FILE.exists():
    st.error(f"⚠️ No se encontró el archivo fijo en: {DEFAULT_FILE}\n\nInclúyelo en la carpeta **data/**.")
    st.stop()

# 👇 Cargar y normalizar los datos
df = read_excel_all(DEFAULT_FILE, VALID_SHEETS)
df = normalize_columns(df)

st.caption(f"Fuente: **{DEFAULT_FILE.name}** · Hojas: {', '.join(VALID_SHEETS)}")


# ================== FUNCIONES AUXILIARES ==================
def available(col, df): return col in df.columns
def options_sorted(series): return sorted([x for x in series.dropna().astype(str).str.strip().unique() if x != ""])
def multiselect_if(col, df, label=None, key=None):
    if available(col, df):
        opts = options_sorted(df[col])
        return st.sidebar.multiselect(label or col, opts, key=key)
    return []
def filter_by_selection(df, col, selected):
    if available(col, df) and selected:
        return df[df[col].isin(selected)]
    return df

# ================== FILTROS ==================
st.sidebar.image("data/betagroup_logo.jpg", width=290)
st.sidebar.markdown("---")
st.sidebar.header("Filtros")
col_btn, _ = st.sidebar.columns([1,1])
with col_btn:
    do_reset = st.button("🔄 Limpiar filtros")
df_f = df.copy()
if not do_reset:
    sel_depto   = multiselect_if("Departamento", df, "Departamento", "depto")
    sel_mpio    = multiselect_if("Municipio", df, "Municipio", "mpio")
    sel_enfoque = multiselect_if("Enfoque Turístico", df, "Enfoque Turístico", "enfoque")
    sel_aspecto = multiselect_if("Aspecto", df, "Aspecto", "aspecto")
    sel_sector  = multiselect_if("Sector", df, "Sector", "sector")
else:
    sel_depto = sel_mpio = sel_enfoque = sel_aspecto = sel_sector = []
df_f = filter_by_selection(df_f, "Departamento", sel_depto)
df_f = filter_by_selection(df_f, "Municipio", sel_mpio)
df_f = filter_by_selection(df_f, "Enfoque Turístico", sel_enfoque)
df_f = filter_by_selection(df_f, "Aspecto", sel_aspecto)
df_f = filter_by_selection(df_f, "Sector", sel_sector)
with st.sidebar.expander("🔎 Búsqueda por texto", expanded=False):
    search_cols = [c for c in ["Nombre", "Actor", "Título", "Descripción"] if available(c, df_f)]
    query = st.text_input("Contiene (min. 2 caracteres)")
    if query and len(query) >= 2 and search_cols:
        mask = pd.Series(False, index=df_f.index)
        for c in search_cols:
            mask = mask | df_f[c].fillna("").astype(str).str.contains(query, case=False, na=False)
        df_f = df_f[mask]
st.sidebar.markdown("---")
st.sidebar.image("data/OIP.webp", width=290)

# ================== KPIs ==================
c1, c2 = st.columns(2)
with c1:
    st.markdown(f'<div class="kpi-card"><p class="kpi-title">Filas filtradas</p><p class="kpi-value">{len(df_f):,}</p></div>', unsafe_allow_html=True)
with c2:
    if available("Departamento", df_f):
        st.markdown(f'<div class="kpi-card"><p class="kpi-title">Departamentos</p><p class="kpi-value">{df_f["Departamento"].nunique():,}</p></div>', unsafe_allow_html=True)
st.divider()

# ================== TABS ==================
tab_resumen, tab_tabla, tab_explorar, tab_barras, tab_sentimientos = st.tabs([
    "📌 Resumen",
    "📋 Tarjetas de Información",
    "🗺️ Mapa Geografico",
    "📊 Barras",
    "💬 Sentimientos"
])

# ... aquí puedes dejar el resto de tu código para resumen, tarjetas, mapa y barras tal como ya lo tienes ...


# ================== CSS PARA AGRANDAR TABS ==================
st.markdown("""
    <style>
    button[data-baseweb="tab"] > div[data-testid="stMarkdownContainer"] p {
        font-size: 18px;   /* tamaño de la letra */
        font-weight: 600;  /* más negrita */
    }
    </style>
""", unsafe_allow_html=True)


# Función de estilo reutilizable (sin highlight_max)
def estilo_tabla(df):
    return (
        df.style
        .set_properties(**{
            "background-color": "#f9fafb",   # gris muy claro
            "color": "#111827",              # texto oscuro
            "border-color": "#e5e7eb",       # bordes suaves
            "border-radius": "8px",          # esquinas redondeadas
            "padding": "6px",                # espacio interno
        })
        .set_table_styles([
            {"selector": "thead th", "props": [("background-color", "#3b82f6"), 
                                               ("color", "white"),
                                               ("font-size", "14px"),
                                               ("padding", "8px"),
                                               ("text-align", "center")]},
            {"selector": "tbody td", "props": [("font-size", "13px"),
                                               ("text-align", "center")]}
        ])
    )

with tab_resumen:
    # Pequeños rankings
    cols = st.columns(2)
    if available("Aspecto", df_f) and not df_f["Aspecto"].dropna().empty:
        top_aspecto = df_f["Aspecto"].value_counts().head(5).reset_index()
        top_aspecto.columns = ["Aspecto", "Conteo"]
        with cols[0]:
            st.subheader("Top 5 Aspectos")
            st.dataframe(estilo_tabla(top_aspecto), use_container_width=True, hide_index=True)

    if available("Enfoque Turístico", df_f) and not df_f["Enfoque Turístico"].dropna().empty:
        top_enfoque = df_f["Enfoque Turístico"].value_counts().head(5).reset_index()
        top_enfoque.columns = ["Enfoque Turístico", "Conteo"]
        with cols[1]:
            st.subheader("Top 5 Enfoques")
            st.dataframe(estilo_tabla(top_enfoque), use_container_width=True, hide_index=True)

# --------- TABLA (AgGrid) ----------
import pandas as pd
import io

with tab_tabla:
    st.subheader("🗂️ Explora los registros en formato tarjetas")

    # --- 📥 Botón para descargar datos filtrados ---
    if len(df_f) > 0:
        # Convertir el DataFrame filtrado a Excel en memoria
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_f.to_excel(writer, index=False, sheet_name='Datos Filtrados')
        buffer.seek(0)

        # Botón de descarga con estilo
        st.download_button(
            label="📥 Descargar datos filtrados (Excel)",
            data=buffer,
            file_name="datos_filtrados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Descarga la información mostrada en las tarjetas.",
            use_container_width=True
        )

    # --- Si no hay resultados ---
    if len(df_f) == 0:
        st.info("No hay filas con los filtros actuales. Ajusta filtros o limpia la búsqueda.")
    else:
        # --- Mostrar tarjetas ---
        for i, row in df_f.head(50).iterrows():  # límite para no sobrecargar
            with st.container():
                st.markdown('<div class="card">', unsafe_allow_html=True)

                # ---- TÍTULO ----
                title = row.get("Título") or row.get("Titulo") or row.get("Nombre") or f"Registro {i}"
                st.markdown(f"<h4>{title}</h4>", unsafe_allow_html=True)

                # ---- BADGES ----
                badges = []
                for b in ["Departamento", "Municipio", "Enfoque Turístico", "Aspecto", "Sector", "Actor"]:
                    v = row.get(b)
                    if v and str(v).strip():
                        badges.append(f'<span class="badge">{b}: {v}</span>')
                if badges:
                    st.markdown(" ".join(badges), unsafe_allow_html=True)

                # ---- DESCRIPCIÓN ----
                desc = (row.get("Descripción") or "").strip()
                if desc:
                    st.markdown("<h5>Descripción</h5>", unsafe_allow_html=True)
                    st.markdown(f"<p>{desc}</p>", unsafe_allow_html=True)

                # ---- APORTE ----
                apInvest = (row.get("Aporte a la Investigación") or "").strip()
                if apInvest:
                    st.markdown("<h5>Aporte a la investigación</h5>", unsafe_allow_html=True)
                    st.markdown(f"<p>{apInvest}</p>", unsafe_allow_html=True)

                # ---- FUENTE ----
                fuente_nombre_cols = ["Fuente", "Fuente / Autor", "Autor", "Autores", "Entidad", "Institución"]
                fuente_url_cols = ["Fuente_URL", "Fuente / Autor_URL", "Autor_URL", "Autores_URL", "Entidad_URL", "Institución_URL"]

                fuente_nombre, fuente_url = "", ""
                for c in fuente_nombre_cols:
                    if c in row and str(row.get(c)).strip() not in ["", "nan", "None"]:
                        fuente_nombre = str(row.get(c)).strip()
                        break

                for c in fuente_url_cols:
                    if c in row and str(row.get(c)).strip() not in ["", "nan", "None"]:
                        fuente_url = str(row.get(c)).strip()
                        break

                if fuente_url and not fuente_url.lower().startswith(("http://", "https://")):
                    fuente_url = "https://" + fuente_url

                if fuente_nombre or fuente_url:
                    st.markdown("<h5>Fuente</h5>", unsafe_allow_html=True)
                    if fuente_url and fuente_nombre:
                        st.markdown(f'<p><a href="{fuente_url}" target="_blank" rel="noopener noreferrer">{fuente_nombre}</a></p>', unsafe_allow_html=True)
                    elif fuente_url:
                        st.markdown(f'<p><a href="{fuente_url}" target="_blank" rel="noopener noreferrer">{fuente_url}</a></p>', unsafe_allow_html=True)
                    elif fuente_nombre:
                        st.markdown(f"<p>{fuente_nombre}</p>", unsafe_allow_html=True)

                st.markdown("</div>", unsafe_allow_html=True)

                # 💚 Línea separadora verde elegante
                st.markdown("""
                <hr style="border: none; border-top: 3px solid #3fb4a1; margin: 18px 0; opacity: 0.6;">
                """, unsafe_allow_html=True)


# --------- EXPLORADOR (Treemap / Sunburst) ----------
# -------------------------------
# -------------------------------
# Cargar GeoJSON de departamentos desde GeoBoundaries
# -------------------------------
@st.cache_data
def cargar_departamentos():
    url_meta = "https://www.geoboundaries.org/api/current/gbOpen/COL/ADM1"
    r = requests.get(url_meta)
    if r.status_code == 200:
        meta = r.json()
        url_geojson = meta["gjDownloadURL"]
        r2 = requests.get(url_geojson)
        if r2.status_code == 200:
            return r2.json()
        else:
            st.error("⚠️ No se pudo descargar el archivo GeoJSON desde GeoBoundaries.")
            return None
    else:
        st.error("⚠️ No se pudo obtener metadatos de GeoBoundaries.")
        return None


# -------------------------------
# Explorador con filtros
# -------------------------------
with tab_explorar:
    geojson_departamentos = cargar_departamentos()

    # -------------------------------
    # Coordenadas aproximadas de municipios
    # -------------------------------
    coords_municipios = {
        # 🔹 Huila
        "Villavieja": [3.2189, -75.2189],
        "Neiva": [2.9386, -75.2819],
        "Garzón": [2.1953, -75.6275],
        "Paicol": [2.4500, -75.7667],
        "Yaguará": [2.6642, -75.5178],
        "San Agustín": [1.8828, -76.2683],
        "Pitalito": [1.8536, -76.0498],

        # 🔹 Tolima
        "Líbano": [4.9211, -75.0622],
        "San Sebastián de Mariquita": [5.1989, -74.8944],
        "Falan": [5.1175, -74.9517],
        "Ibagué": [4.4389, -75.2322],
        "Honda": [5.0713, -74.6949],
        "Armero": [5.0300, -74.9000],
        "Prado": [3.7500, -74.9167],

        # 🔹 Putumayo
        "Puerto Asís": [0.5052, -76.4951],
        "Orito": [0.6781, -76.8723],
        "Puerto Caicedo": [0.6953, -76.6044],
        "Valle del Guamuez": [0.4519, -76.9292],
        "Villagarzón": [0.9892, -76.6279],
        "Mocoa": [1.1474, -76.6473],
        "Sibundoy": [1.2081, -76.9220],
        "Colón": [1.1900, -76.9740],
        "Santiago": [1.1461, -77.0031],
        "San Francisco": [1.1761, -76.8789],

        # 🔹 Caquetá
        "San Vicente del Caguán": [2.1167, -74.7667],
        "Doncello": [1.6789, -75.2806],
        "Florencia": [1.6144, -75.6062],
        "San José del Fragua": [1.3300, -75.9700],
        "Belén de los Andaquies": [1.4167, -75.8667],
    }

    # --- Relación de municipios por departamento ---
    municipios_por_departamento = {
        "Huila": ["Villavieja", "Neiva", "Garzón", "Paicol", "Yaguará", "San Agustín", "Pitalito"],
        "Tolima": ["Líbano", "San Sebastián de Mariquita", "Falan", "Ibagué", "Honda", "Armero", "Prado"],
        "Putumayo": ["Puerto Asís", "Orito", "Puerto Caicedo", "Valle del Guamuez", "Villagarzón", "Mocoa", "Sibundoy", "Colón", "Santiago", "San Francisco"],
        "Caquetá": ["San Vicente del Caguán", "Doncello", "Florencia", "San José del Fragua", "Belén de los Andaquies"],
    }

    with st.container():
        st.subheader("🗺️ Explorador geográfico con filtros")
        st.caption("Filtra por Departamento, Municipio, Aspecto, Enfoque o Sector y visualiza los resultados en el mapa.")

        # --- Columnas disponibles dinámicamente ---
        dims = [d for d in ["Departamento", "Municipio", "Aspecto", "Enfoque Turístico", "Sector"] if d in df_f.columns]

        if len(dims) == 0:
            st.info("⚠️ No se encuentran columnas categóricas para filtrar.")
        else:
            # Crear filtros dinámicos
            filtros = {}
            for dim in dims:
                valores = sorted(df_f[dim].dropna().unique())
                seleccion = st.multiselect(f"📍 Filtrar por {dim}:", valores, default=valores)
                filtros[dim] = seleccion

            # Aplicar filtros
            df_filtrado = df_f.copy()
            for dim, seleccion in filtros.items():
                if seleccion:
                    df_filtrado = df_filtrado[df_filtrado[dim].isin(seleccion)]

            # Layout en dos columnas
            col1, col2 = st.columns([2, 2])

            # --- Resumen en tabla ---
            with col1:
                st.markdown("### 📊 Resumen filtrado")
                if len(df_filtrado) == 0:
                    st.info("No hay registros con los filtros seleccionados.")
                else:
                    resumen = df_filtrado.groupby(dims).size().reset_index(name="Conteo")
                    st.dataframe(resumen, use_container_width=True)

            # --- Mapa geográfico ---
            with col2:
                st.markdown("### 🗺️ Mapa interactivo")
                if len(df_filtrado) == 0:
                    st.caption("No hay datos para mostrar en el mapa.")
                else:
                    departamentos = df_filtrado["Departamento"].unique()

                    # Crear mapa
                    m = folium.Map(location=[2.5, -75.0], zoom_start=6, tiles="cartodbpositron")

                    # --- Colores para departamentos filtrados ---
                    colores = ["#FF5733", "#33FF57", "#3357FF", "#F1C40F", "#9B59B6", "#E67E22", "#1ABC9C"]
                    color_map = {depto: colores[i % len(colores)] for i, depto in enumerate(departamentos)}

                    # Dibujar departamentos y municipios
                    if geojson_departamentos:
                        for feature in geojson_departamentos["features"]:
                            nombre_depto = feature["properties"]["shapeName"]
                            if nombre_depto in departamentos:
                                folium.GeoJson(
                                    feature,
                                    name=nombre_depto,
                                    style_function=lambda f, nombre=nombre_depto: {
                                        "fillColor": color_map[nombre],
                                        "color": "black",
                                        "weight": 2,
                                        "fillOpacity": 0.5,
                                    },
                                    tooltip=folium.GeoJsonTooltip(fields=["shapeName"], aliases=["Departamento:"]),
                                ).add_to(m)

                                # --- Marcar municipios de ese departamento con el mismo color ---
                                municipios = municipios_por_departamento.get(nombre_depto, [])
                                for municipio in municipios:
                                    coords_mun = coords_municipios.get(municipio)
                                    if coords_mun:
                                        folium.Marker(
                                            location=coords_mun,
                                            popup=f"<b>{municipio}</b><br>Departamento: {nombre_depto}",
                                            tooltip=f"{municipio} ({nombre_depto})",
                                            icon=folium.Icon(color="blue", icon_color=color_map[nombre_depto], icon="", prefix="fa"),
                                        ).add_to(m)

                    # Mostrar mapa
                    st_folium(m, width=900, height=600)

# --------- BARRAS DINÁMICAS ----------
with tab_barras:
    st.subheader("📊 Comparación por categorías")

    cols = [c for c in ["Aspecto", "Enfoque Turístico", "Municipio", "Departamento", "Sector"] if available(c, df_f)]
    
    if not cols:
        st.info("No hay columnas categóricas disponibles para graficar.")
    else:
        col_sel = st.selectbox("📍 Selecciona categoría", cols, index=0)
        top_n = st.slider("🔝 Top N", 5, 50, 20, step=5)

        s = df_f[col_sel].dropna().astype(str).str.strip()
        s = s[s != ""]
        
        if s.empty:
            st.info("⚠️ No hay datos válidos para la categoría seleccionada.")
        else:
            vc = s.value_counts().head(top_n).reset_index()
            vc.columns = [col_sel, "Conteo"]

            fig = px.bar(
                vc,
                x="Conteo",
                y=col_sel,
                orientation="h",
                text="Conteo",
                color="Conteo",
                color_continuous_scale="plasma"  # gradiente bonito
            )

            fig.update_traces(
                texttemplate="%{text}",
                textposition="outside",
                marker=dict(line=dict(width=0.5, color="white"))
            )

            fig.update_layout(
                yaxis={"categoryorder": "total ascending"},
                height=600,
                margin=dict(l=10, r=10, t=30, b=10),
                xaxis_title="Número de registros",
                yaxis_title="",
                plot_bgcolor="rgba(0,0,0,0)",  # fondo transparente
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(size=13)
            )

            st.plotly_chart(fig, use_container_width=True)
            # --------- NUEVA PESTAÑA: ANÁLISIS DE SENTIMIENTOS ----------
with tab_sentimientos:
    st.subheader("💬 Análisis de Sentimientos Identificados")

    if not available("Sentimiento identificado", df_f):
        st.warning("⚠️ No se encontró la columna 'Sentimiento identificado' en los datos.")
    else:
        # Normalizar texto de la columna
        df_f["Sentimiento identificado"] = (
            df_f["Sentimiento identificado"]
            .astype(str)
            .str.strip()
            .str.capitalize()
        )

        # Contar los sentimientos
        sentiment_counts = (
            df_f["Sentimiento identificado"]
            .value_counts()
            .reset_index()
        )
        sentiment_counts.columns = ["Sentimiento identificado", "Conteo"]

        # Mostrar tabla resumen
        st.markdown("### 📋 Distribución de sentimientos")
        st.dataframe(estilo_tabla(sentiment_counts), use_container_width=True, hide_index=True)

        # Colores personalizados según sentimiento
        sentiment_colors = {
            "Muy positivo": "#2ECC71",   # Verde fuerte
            "Positivo": "#58D68D",       # Verde claro
            "Neutro": "#B2BABB",         # Gris
            "Negativo": "#E67E22",       # Naranja
            "Muy negativo": "#E74C3C",   # Rojo
        }

        # Crear gráfico circular
        fig_pie = px.pie(
            sentiment_counts,
            names="Sentimiento identificado",
            values="Conteo",
            color="Sentimiento identificado",
            color_discrete_map=sentiment_colors,
            hole=0.3,
            title="Distribución porcentual de sentimientos"
        )
        fig_pie.update_traces(
            textposition="inside",
            textinfo="percent+label",
            pull=[0.05] * len(sentiment_counts)
        )
        st.plotly_chart(fig_pie, use_container_width=True)

        # Mostrar registros por sentimiento
        st.markdown("### 🔍 Registros agrupados por sentimiento")
        sentimiento_sel = st.selectbox(
            "Selecciona un sentimiento para explorar ejemplos:",
            sentiment_counts["Sentimiento identificado"]
        )

        df_sel = df_f[df_f["Sentimiento identificado"] == sentimiento_sel]
        if df_sel.empty:
            st.info("No hay registros con este sentimiento.")
        else:
            for i, row in df_sel.head(20).iterrows():
                st.markdown(
                    f"**• {row.get('Título', 'Sin título')}** — "
                    f"{row.get('Descripción', '')[:200]}..."
                )
